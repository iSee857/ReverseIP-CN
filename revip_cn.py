import re
import sys
import socket
import random
import getopt
import requests
import openpyxl
import time
from urllib.parse import urlparse
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor


HIGHLIGHT_FILL = PatternFill(start_color='FFFF00', fill_type='solid')
HEADER_FILL = PatternFill(start_color='DDDDDD', fill_type='solid')
VERSION = "V2.1"
AUTHOR = "iSee857"

def print_banner():
    banner = f"""
██████╗ ███████╗███████╗██████╗ ███████╗██████╗ ███████╗
██╔══██╗██╔════╝██╔════╝██╔══██╗██╔════╝██╔══██╗██╔════╝
██████╔╝█████╗  █████╗  ██║  ██║█████╗  ██████╔╝███████╗
██╔══██╗██╔══╝  ██╔══╝  ██║  ██║██╔══╝  ██╔══██╗╚════██║
██║  ██║███████╗██║     ██████╔╝███████╗██║  ██║███████║
╚═╝  ╚═╝╚══════╝╚═╝     ╚═════╝ ╚══════╝╚═╝  ╚═╝╚══════╝
Reverse IP Lookup Tool {VERSION}
Author: {AUTHOR}
"""
    print(banner)

def clean_target(target):
    """智能清洗输入目标"""
    try:
        target = target.strip(" '\"")
        if not target.startswith(('http://', 'https://')):
            target = f'http://{target}'
        
        parsed = urlparse(target)
        hostname = parsed.hostname
        if not hostname:
            return None
        
        return socket.gethostbyname(hostname)
    except Exception as e:
        print(f"解析失败: {str(e)}")
        return None

def user_agents():
    """国内主流浏览器User-Agent"""
    return [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:83.0) Gecko/20100101 Firefox/83.0",
        "Mozilla/5.0 (Linux; Android 10; M2007J3SC) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Mobile Safari/537.36",
        "Mozilla/5.0 (iPhone; CPU iPhone OS 14_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.1 Mobile/15E148 Safari/604.1"
    ]

def fetch_domains_cn(ip):
    """使用国内接口查询IP关联域名"""
    headers = {
        'User-Agent': random.choice(user_agents()),
        'Referer': 'https://site.ip138.com/'
    }
    """国内接口可配"""
    apis = [
        {
            'url': f'https://site.ip138.com/{ip}/',
            'method': 'regex',
            'pattern': r'<li><span class="date">.*?</span><a href="/(.*?)/" target="_blank">'
        },
        #站点关闭
        # {
        #     'url': f'https://api.webscan.cc/?query={ip}',
        #     'method': 'json',
        #     'field': 'domain'
        # }
    ]
    
    domains = []
    for api in apis:
        try:
            session = requests.Session()
            session.trust_env = False
            response = session.get(
                api['url'],
                headers=headers,
                timeout=15,
                proxies={'http': None, 'https': None}
            )
            
            if response.status_code != 200:
                continue
                
            if api['method'] == 'regex':
                matches = re.findall(api['pattern'], response.text)
                cleaned = [m.strip() for m in matches if m.strip()]
                domains.extend(cleaned)
                
            elif api['method'] == 'json':
                data = response.json()
                if isinstance(data, list):
                    valid = [str(d.get(api['field'], '')).strip() for d in data]
                    domains.extend([v for v in valid if v])
            
            # 去重并限制最大数量防止站点重定向
            domains = list(set(domains))[:50]
            time.sleep(random.uniform(1, 2))
            
        except Exception as e:
            print(f"接口 {api['url']} 查询失败: {str(e)}")
            continue
            
    return domains

def process_target(target):
    """处理目标"""
    ip = clean_target(target)
    if not ip:
        print(f"\n❌ 目标解析失败: {target}")
        print("-" * 50)
        return (target, None)
    
    domains = fetch_domains_cn(ip)
    
    original_host = urlparse(target).hostname or target.split('//')[-1].split('/')[0]
    highlighted_domains = [
        f"\033[93m{d}*\033[0m" if (ip in d or original_host in d) else d 
        for d in domains if isinstance(d, str)
    ]
    
    ip_display = f"\033[92m{ip}\033[0m" if domains else ip
    
    print(f"\n► 原始输入: \033[94m{target}\033[0m")
    print(f"► 解析IP : {ip_display}")
    print(f"► 关联域名: {len(domains)} 个")
    
    if domains:
        print("  " + "\n  ".join(highlighted_domains))
    else:
        print("  未找到关联域名")
    
    print("-" * 50) 
    
    return (target, {"ip": ip, "domains": domains})

def export_results(results, filename):
    """导出Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "反查结果"
    
    headers = ["原始输入", "IP地址", "关联域名"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).fill = HEADER_FILL
    
    row_idx = 2
    for target, data in results.items():
        if not data:
            ws.append([target, None, "解析失败"])
            continue
            
        domains = data['domains']
        original_host = urlparse(target).hostname
        
        highlight = any(
            (data['ip'] in d) or 
            (original_host and original_host in d)
            for d in domains if isinstance(d, str)
        )
        
        row = [
            target,
            data['ip'],
            "\n".join(domains) if domains else "无结果"
        ]
        ws.append(row)
        
        if highlight:
            ws.cell(row=row_idx, column=3).fill = HIGHLIGHT_FILL
        
        row_idx += 1
    
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    
    wb.save(filename)
    print(f"\n✅ 结果已保存到: {filename}")

def main(argv):
    print_banner()
    targets = []
    output = "results.xlsx"

    try:
        opts, args = getopt.getopt(argv, "hu:l:o:", ["help", "url=", "list=", "output="])
    except getopt.GetoptError:
        print("参数错误！使用 -h 查看帮助")
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            print(f"Usage: {sys.argv[0]} [-u URL/IP] [-l FILE] [-o FILE]")
            sys.exit()
        elif opt in ("-u", "--url"):
            targets.append(arg)
        elif opt in ("-l", "--list"):
            try:
                with open(arg, 'r') as f:
                    targets.extend(line.strip() for line in f if line.strip())
            except FileNotFoundError:
                print(f"文件不存在: {arg}")
                sys.exit(1)
        elif opt in ("-o", "--output"):
            output = arg

    if not targets:
        print("请指定目标(-u/-l)")
        sys.exit(1)

    print(f"\n🔍 开始处理 {len(targets)} 个目标...")
    
    results = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(process_target, t) for t in targets]
        for future in futures:
            target, data = future.result()
            results[target] = data
    
    export_results(results, output)

if __name__ == "__main__":
    main(sys.argv[1:])
